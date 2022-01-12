#Librerias necesarias para que funcione el programa
import requests
import xlsxwriter
import json
import openpyxl
from openpyxl.styles import Font

#idTorneo es el identificador del torneo el cual debe de ser anexado a la variable url;
idTorneo = "1QnbUgkM"
url = 'https://lichess.org/api/tournament/'+idTorneo
#Esta variable obtendra los resultados del torneo
r = requests.get(url+'/results', allow_redirects=True)
#Esta variable obtendra la información completa del torneo
rn = requests.get(url, allow_redirects=True)
#Se guarda el Json para obtener el nombre
infTor = json.loads(rn.content)
#Obtenemos el nombre completo del torneo
nombreTorneo = infTor['fullName']
#Tipo de torneo, 1 para puntos normales, 2 para puntos modificados
bonificacion = 2
#Creamos los documentos con la información requerida para crear nuestra lista de rendimiento.
open('C:/creaListaDerendimiento/results', 'wb').write(r.content)
f = open('C:/creaListaDerendimiento/results', "r")
g = open('C:/creaListaDerendimiento/filter', "w+")
#FIltramos unicamente los miembros de nuestro equipo y los escribimos en nuestro archivo.
for x in f:
    y = json.loads(x)
    #Especificamos de cual equipo queremos obtener sus jugadores con su respectiva información.
    if y["team"] == "caballo-negro-chess-club":
        g.write(x)
f.close()
g.close()
#Creamos una lista previa en excel
workbook = xlsxwriter.Workbook('C:/creaListaDerendimiento/preview.xlsx')
worksheet = workbook.add_worksheet()
row = 0
col = 0
g = open('C:/creaListaDerendimiento/filter', "r")
for x in g:
    y = json.loads(x)
    worksheet.write(row, col, str(y["username"]))
    worksheet.write(row, col + 1, str(y["score"]))
    row += 1
g.close()
workbook.close()
#Acomodar los puntos realizados y en base de de ellos sumarlos en el excel de forma acumulativa.
torneoActual = openpyxl.load_workbook('C:/creaListaDerendimiento/preview.xlsx')
rendimientoActual = openpyxl.load_workbook('C:/creaListaDerendimiento/final.xlsx')
sheet_rangesTA = torneoActual['Sheet1']
sheet_rangesRA = rendimientoActual['Sheet1']
#La funcion .max_row encuentra exactamente el numero de filas, si excel tiene n con esta funcion devolveremos n
#El numero de columas del documento lo obtendremos con cuantos torneos han sido contabilizados ya
columnaMax = sheet_rangesRA['B4'].value + 1
caracterColumna = ""+chr(columnaMax + 97)
columnaTotal = chr(columnaMax + 97 + 1)
#Limpiamos residuos del total
if columnaMax !=1:
    for a in range(sheet_rangesRA.max_row - 5):
        sheet_rangesRA[caracterColumna + str(a + 5)] = ""
#Al agregar un nuevo torneo este debe de ser contabilizado
sheet_rangesRA['B4'].value = sheet_rangesRA['B4'].value + 1
#Comenzamos a traspasar la información del nuevo torneo
if columnaMax == 1:
    for x in range(sheet_rangesTA.max_row):
        sheet_rangesRA['A' + str(x + 6)].value = sheet_rangesTA['A' + str(x+1)].value
        sheet_rangesRA["{}{}".format(caracterColumna, str(x + 6))].value = (int(sheet_rangesTA["{}{}".format('B', str(x + 1))].value))*bonificacion
else:
    for n in range(sheet_rangesRA.max_row-5):
        for m in range(sheet_rangesTA.max_row):
            if sheet_rangesRA['A' + str(n + 6)].value == sheet_rangesTA['A' + str(m + 1)].value:
                #Para que se guarden los numeros como numeros
                sheet_rangesRA["{}{}".format(caracterColumna, str(n+6))].value = (int(sheet_rangesTA["{}{}".format('B', str(m+1))].value))*bonificacion
                break
    for m in range(sheet_rangesTA.max_row):
        auxRep = 0
        for n in range(sheet_rangesRA.max_row - 5):
            if sheet_rangesRA['A' + str(n + 6)].value != sheet_rangesTA['A' + str(m + 1)].value:
                auxRep = auxRep + 1
        if auxRep >= sheet_rangesRA.max_row - 5:
            sheet_rangesRA['A' + str(sheet_rangesRA.max_row+1)].value = sheet_rangesTA['A' + str(m + 1)].value
            sheet_rangesRA["{}{}".format(caracterColumna, str(sheet_rangesRA.max_row))].value = (int(sheet_rangesTA["{}{}".format('B', str(m + 1))].value))*bonificacion
for a in range(sheet_rangesRA.max_row - 5):
    if sheet_rangesRA['A' + str(a+6)].value != None:
        sheet_rangesRA[columnaTotal + str(a+6)] = "=SUM(B"+str(a+6)+":"+caracterColumna+str(a+6)+")"
sheet_rangesRA[columnaTotal + str(5)].value = "Total"
sheet_rangesRA[caracterColumna + str(5)].value = nombreTorneo
if bonificacion == 2:
    sheet_rangesRA[caracterColumna + str(5)].font = Font(color="3b83bd")
rendimientoActual.save("C:/creaListaDerendimiento/final.xlsx")





