#Codigo propio
import  openpyxl
torneoActual = openpyxl.load_workbook('C:/creaListaDerendimiento/preview.xlsx')
rendimientoActual = openpyxl.load_workbook('C:/creaListaDerendimiento/final.xlsx')
sheet_rangesTA = torneoActual['Sheet1']
sheet_rangesRA = rendimientoActual['Sheet1']
for x in range(1, sheet_rangesTA.max_row + 1):
    if sheet_rangesRA['A' + str(x + 4)].value is None:
        sheet_rangesRA['A' + str(x + 4)].value = sheet_rangesTA['A' + str(x)].value
    else:
        for y in range(1, sheet_rangesTA.max_row + 1):
            if sheet_rangesRA['A'+str(x+4)].value == sheet_rangesTA['A' + str(y)].value:
                print(sheet_rangesRA['A'+str(x+4)].value)
                break
            else:
                aux =+1
            if(aux == sheet_rangesTA.max_row + 1):
                sheet_rangesRA['A'+str(sheet_rangesTA.max_row + 1)].value = sheet_rangesTA['A' + str(y)].value
#for x in range(1, sheet_rangesTA.max_row + 1):
#    if sheet_rangesRA['A'+str(x+4)].value is None:
#        sheet_rangesRA['A' + str(x + 4)].value = sheet_rangesTA['A' + str(x)].value
#    elif sheet_rangesRA['A'+str(x+4)].value == sheet_rangesTA['A' + str(x)].value:
        #Sumar puntos
rendimientoActual.save("C:/creaListaDerendimiento/final.xlsx")
#print(sheet_rangesTA['A1'].value)
#print(sheet_rangesRA['A4'].value)
